import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

sheet_name_mapping = {
    "VENKATARAMAN_3980807": "Sadhana Venkat",
    "Venkataraman_3657521": "Sowmya Venkat",
    "V_2654323":            "Ashwin Ram",
    "Ramachandran_852890":  "Venkataraman R"
}

columns_to_drop = ["broker_code"]


def to_number(x):
    if pd.isna(x):
        return None
    s = str(x).replace(",", "").strip()
    if s in ["-", ""]:
        return None
    try:
        return float(s)
    except Exception:
        return None


def calculate_holding_period(invest_date, today=None):
    today = today or datetime.today()
    if pd.isna(invest_date):
        return None
    if isinstance(invest_date, str):
        invest_date = pd.to_datetime(invest_date, errors="coerce")
    if invest_date is pd.NaT:
        return None
    delta = today - invest_date
    years  = delta.days // 365
    months = (delta.days % 365) // 30
    return f"{years}y {months}m"


def calculate_ltcg_stcg(row, today=None):
    today      = today or datetime.today()
    asset_class = str(row.get("Asset Class", "")).lower()
    invest_date = row.get("Investment Date")
    if pd.isna(invest_date):
        return "NA"
    invest_date = pd.to_datetime(invest_date, errors="coerce")
    if invest_date is pd.NaT:
        return "NA"
    holding_days = (today - invest_date).days
    if "equity" in asset_class:
        return "LTCG" if holding_days > 365 else "STCG"
    elif "debt" in asset_class:
        return "LTCG" if holding_days > 1095 else "STCG"
    return "NA"


def clean_indmoney_report(input_path=None, output_path=None):
    project_root = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))

    if input_path is None:
        input_path  = os.path.join(project_root, "database", "data", "raw", "indmoney", "latest.xlsx")
    if output_path is None:
        output_path = os.path.join(project_root, "database", "data", "processed", "cleaned_indmoney_family_report.xlsx")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Read all sheets
    sheets = pd.read_excel(input_path, sheet_name=None, skiprows=5)
    sheets = {k: v for k, v in sheets.items() if k in sheet_name_mapping}

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():

            df = df.dropna(axis=1, how="all")

            if len(df.columns) > 0 and str(df.columns[0]).startswith("source_holding_id"):
                df = df.iloc[:, 1:]

            if "First Name" in df.columns:
                df = df.rename(columns={"First Name": "Name"})

            df = df.drop(columns=[c for c in columns_to_drop if c in df.columns])

            if "Asset Type" in df.columns:
                # Do not drop stocks anymore, map them to US Stock or IND Equity based on category if needed later
                pass

            new_sheet_name = sheet_name_mapping[sheet_name]

            if "Name" in df.columns:
                df["Name"] = new_sheet_name

            if "Asset Type" in df.columns:
                df["Asset Type"] = (
                    df["Asset Type"]
                    .astype(str)
                    .str.replace("_", " ", regex=False)
                    .str.title()
                    .replace({
                        "Mf":       "Mutual Fund",
                        "Sa":       "Savings Account",
                        "Us Stock": "US Stock",
                        "Fd":       "Fixed Deposit"
                    })
                )

            numeric_cols = [
                "Total Units",
                "Invested Amount",
                "Market Value",
                "Holding (%)",
                "Total Gain/Loss (INR)",
                "Total Gain/Loss (%)",
                "XIRR (%)"
            ]
            for c in numeric_cols:
                if c in df.columns:
                    df[c] = df[c].apply(to_number)

            if "Investment Code" in df.columns:
                df["investment_code"] = df["Investment Code"].astype(str)
                df = df.drop(columns=["Investment Code"])

            direct_col = next((c for c in df.columns if "Direct/Regular" in c), None)
            broker_col = next((c for c in df.columns if "Broker" in c), None)
            if direct_col and broker_col:
                df["Direct/Regular - Broker"] = df[direct_col].astype(str) + " - " + df[broker_col].astype(str)
            elif direct_col:
                df["Direct/Regular - Broker"] = df[direct_col].astype(str)
            elif broker_col:
                df["Direct/Regular - Broker"] = df[broker_col].astype(str)

            if "Investment Date" in df.columns:
                df["Holding Period"] = df["Investment Date"].apply(calculate_holding_period)

            df["LTCG/STCG"] = df.apply(calculate_ltcg_stcg, axis=1)

            if "Expense Ratio" in df.columns:
                df["Expense Ratio"] = df["Expense Ratio"].apply(lambda x: x / 100 if x and x > 1 else x)

            if "AMC Name" in df.columns:
                df = df.drop(columns=["AMC Name"])

            desired_order = [
                "Name", "Asset Type", "Asset Class", "Category", "Investment",
                "Total Units", "Invested Amount", "Market Value",
                "Total Gain/Loss (INR)", "Total Gain/Loss (%)", "XIRR (%)",
                "Investment Date", "Holding Period", "LTCG/STCG",
                "Expense Ratio", "Direct/Regular - Broker", "investment_code"
            ]
            df = df[[c for c in desired_order if c in df.columns]]
            df.to_excel(writer, sheet_name=new_sheet_name, index=False)

    # Formatting
    wb     = load_workbook(output_path)
    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        for col in ws.columns:
            max_len    = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_len + 2

        headers   = {c.value: c.column_letter for c in ws[1]}
        gain_cols = ["Total Gain/Loss (INR)", "Total Gain/Loss (%)"]

        for col_name in gain_cols:
            if col_name in headers:
                col_letter = headers[col_name]
                rng        = f"{col_letter}2:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=green_fill))
                ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",    formula=["0"], fill=red_fill))
                for cell in ws[col_letter][1:]:
                    cell.font = Font(bold=True)

        currency_cols = ["Invested Amount", "Market Value", "Total Gain/Loss (INR)"]
        for col_name in currency_cols:
            if col_name in headers:
                col_letter = headers[col_name]
                for cell in ws[col_letter][1:]:
                    cell.number_format = '₹#,##0.00'

        if "Expense Ratio" in headers:
            col_letter = headers["Expense Ratio"]
            for cell in ws[col_letter][1:]:
                if cell.value is not None:
                    cell.number_format = '0.00%'

    wb.save(output_path)
    print(f"Cleaned file saved to: {output_path}")
    return output_path


if __name__ == "__main__":
    clean_indmoney_report()